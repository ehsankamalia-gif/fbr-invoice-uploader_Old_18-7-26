from typing import List, Dict, Tuple, Any
from pathlib import Path

from openpyxl import load_workbook

class ExcelProcessingService:
    def get_sheet_names(self, file_path: str) -> List[str]:
        """Reads an Excel file and returns a list of all sheet names."""
        try:
            path = Path(file_path)
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
            try:
                return list(wb.sheetnames)
            finally:
                wb.close()
        except Exception as e:
            # Consider logging the error
            raise ValueError(f"Could not read sheet names from {file_path}: {e}")

    def read_excel(self, file_path: str, sheet_name: str | None = None) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Reads a specific sheet from an Excel file and returns data and headers."""
        try:
            path = Path(file_path)
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
            try:
                if sheet_name:
                    if sheet_name not in wb.sheetnames:
                        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}")
                    ws = wb[sheet_name]
                else:
                    ws = wb[wb.sheetnames[0]]

                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    return [], []

                raw_headers = [str(col).strip() if col is not None else "" for col in rows[0]]
                sanitized_headers = [h.lower().replace(" ", "_") for h in raw_headers]

                used_indexes: List[int] = []
                headers: List[str] = []
                for idx, header in enumerate(sanitized_headers):
                    if header:
                        used_indexes.append(idx)
                        headers.append(header)

                data: List[Dict[str, Any]] = []
                for row in rows[1:]:
                    if row is None:
                        continue
                    if all(cell is None or str(cell).strip() == "" for cell in row):
                        continue

                    record: Dict[str, Any] = {}
                    for header, col_idx in zip(headers, used_indexes):
                        record[header] = row[col_idx] if col_idx < len(row) else None
                    data.append(record)

                return data, headers
            finally:
                wb.close()
        except Exception as e:
            raise ValueError(f"Failed to read Excel sheet '{sheet_name or 'default'}': {e}")

    def validate_data(self, data: List[Dict[str, Any]], headers: List[str]) -> Dict[str, Any]:
        """Validates the data, looking for a 'phone' or 'number' column."""
        phone_column = None
        keywords = ['phone', 'number', 'cell', 'mobile', 'contact', 'wa']
        
        # Try exact matches first
        for col in headers:
            if col.lower() in keywords:
                phone_column = col
                break
        
        # If no exact match, try keyword containment
        if not phone_column:
            for col in headers:
                if any(kw in col.lower() for kw in keywords):
                    phone_column = col
                    break
        
        if not phone_column:
            return {
                "success": False,
                "error": "Missing required column: 'phone', 'number', 'cell', or 'mobile'.",
                "valid_count": 0,
                "invalid_count": len(data),
                "valid_data": []
            }

        valid_data = []
        invalid_count = 0
        
        for row in data:
            phone_val = row.get(phone_column)
            if phone_val and isinstance(phone_val, (str, int)):
                # Basic normalization of phone number
                phone_str = str(phone_val).strip()
                if phone_str.isdigit() and len(phone_str) >= 10:
                    valid_data.append(row)
                    continue
            
            invalid_count += 1

        return {
            "success": True,
            "valid_count": len(valid_data),
            "invalid_count": invalid_count,
            "valid_data": valid_data
        }

excel_processing_service = ExcelProcessingService()
