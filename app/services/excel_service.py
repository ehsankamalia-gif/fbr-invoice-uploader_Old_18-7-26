from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

def parse_recipients(file_path: str):
    path = Path(file_path)
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], []

        raw_headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        headers = [h for h in raw_headers if h]

        records: List[Dict[str, Any]] = []
        for row in rows[1:]:
            if row is None:
                continue
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            record: Dict[str, Any] = {}
            for idx, header in enumerate(headers):
                record[header] = row[idx] if idx < len(row) else None
            records.append(record)

        return records, headers
    finally:
        wb.close()
